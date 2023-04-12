import time,sys,os
import win32gui, win32con,win32api
import win32clipboard as wc
import uiautomation as auto
import openpyxl
import configparser
import traceback
import easygui
import re
import csv

表格头 = ['订单编号','调货员/现货','货号','尺码','调货价格','淘宝价格','商家备忘','点数','扣除','利润','备注','运费','店铺名称','支付宝','昵称']

# 获取表格实例
def create_excel_xlsx(path, sheet_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.save(path)

# 追加方式写入表格
def write_excel_xlsx_append(path, value, truncate_sheet=False):
    if not os.path.exists(path):
        create_excel_xlsx(path, 'Sheet1')
        data = openpyxl.load_workbook(path)
        # 取第一张表
        sheetnames = data.sheetnames
        sheet = data[sheetnames[0]]
        for i in range(0, len(表格头)):
            sheet.cell(row=1, column=i + 1, value=str(表格头[i]))
        data.save(path)
    data = openpyxl.load_workbook(path)
    # 取第一张表
    sheetnames = data.sheetnames
    sheet = data[sheetnames[0]]
    sheet = data.active
    if(truncate_sheet): #truncate_sheet为True，覆盖原表中的数据
        startrows = 0
    else:
        startrows = sheet.max_row  # 获得行数
    index = len(value)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=startrows + i + 1, column=j + 1, value=str(value[i][j]))
    data.save(path)

# 读取表格
def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

# 对配置文件的处理的类
class dispose_ini:
    """
    封装一个类，进行ini文件的常用操作
    """
    def __init__(self, filepath):
        self._path = filepath
        self.config = configparser.ConfigParser()  # 实例化解析对象
        self.config.read(filepath)  # 读文件
 
    def get_sections(self):
        """
        获取ini文件所有的块，返回为list
        """
        sect = self.config.sections()
        return sect
 
    def get_options(self, sec):
        """
        获取ini文件指定块的项
        :param sec: 传入的块名
        :return: 返回指定块的项（列表形式）
        """
        return self.config.options(sec)
 
    def get_items(self, sec):
        """
        获取指定section的所有键值对
        :param sec: 传入的块名
        :return: section的所有键值对（元组形式）
        """
        return self.config.items(sec)
 
    def get_option(self, sec, opt):
        """
        :param sec: 传入的块名
        :param opt: 传入项
        :return: 返回项的值(string类型)
        """
        return self.config.get(sec, opt)
 
    def write_(self):
        """ 将修改后写入文件 """
        with open(self._path, 'w') as fp:
            self.config.write(fp)
 
    def add_section(self, sec):
        """
        为ini文件添加新的section, 如果section 已经存在则抛出异常
        :param sec: 传入的块名
        :return: None
        """
        self.config.add_section(sec)
        self.write_()
 
    def set_option(self, sec, opt, value):
        """
        对指定section下的某个option赋值
        :param sec: 传入的块名
        :param opt: 传入的项名
        :param value: 传入的值
        :return:  None
        """
        self.config.set(sec, opt, value)
        self.write_()  # 写入文件
 
    def remove_sec(self, sec):
        """
        删除某个section
        :param sec: 传入的块名
        :return: bool
        """
        self.config.remove_section(sec)
        self.write_()  # 写入文件
 
    def remove_opt(self, sec, opt):
        """
        删除某个section下的某个option
        :param sec: 传入的块名
        :param opt: 传入的项名
        :return: bool
        """
        self.config.remove_option(sec, opt)
        self.write_()  # 写入文件

class WxParam:
    SYS_TEXT_HEIGHT = 33
    TIME_TEXT_HEIGHT = 34
    RECALL_TEXT_HEIGHT = 45
    CHAT_TEXT_HEIGHT = 52
    CHAT_IMG_HEIGHT = 117
    SpecialTypes = ['[文件]', '[图片]', '[视频]', '[音乐]', '[链接]']

class WxUtils:
    def SplitMessage(MsgItem):
        auto.SetGlobalSearchTimeout(0)
        MsgItemName = MsgItem.Name
        if MsgItem.BoundingRectangle.height() == WxParam.SYS_TEXT_HEIGHT:
            Msg = ('SYS', MsgItemName, ''.join([str(i) for i in MsgItem.GetRuntimeId()]))
        elif MsgItem.BoundingRectangle.height() == WxParam.TIME_TEXT_HEIGHT:
            Msg = ('Time', MsgItemName, ''.join([str(i) for i in MsgItem.GetRuntimeId()]))
        elif MsgItem.BoundingRectangle.height() == WxParam.RECALL_TEXT_HEIGHT:
            if '撤回' in MsgItemName:
                Msg = ('Recall', MsgItemName, ''.join([str(i) for i in MsgItem.GetRuntimeId()]))
            else:
                Msg = ('SYS', MsgItemName, ''.join([str(i) for i in MsgItem.GetRuntimeId()]))
        else:
            Index = 1
            User = MsgItem.ButtonControl(foundIndex=Index)
            try:
                while True:
                    if User.Name == '':
                        Index += 1
                        User = MsgItem.ButtonControl(foundIndex=Index)
                    else:
                        break
                Msg = (User.Name, MsgItemName, ''.join([str(i) for i in MsgItem.GetRuntimeId()]))
            except Exception as e:
                Msg = ('SYS', MsgItemName, ''.join([str(i) for i in MsgItem.GetRuntimeId()]))
        auto.SetGlobalSearchTimeout(10.0)
        return Msg
    
    def SetClipboard(data, dtype='text'):
        '''复制文本信息或图片到剪贴板
        data : 要复制的内容，str 或 Image 图像'''
        if dtype.upper() == 'TEXT':
            type_data = win32con.CF_UNICODETEXT
        elif dtype.upper() == 'IMAGE':
            from io import BytesIO
            type_data = win32con.CF_DIB
            output = BytesIO()
            data.save(output, 'BMP')
            data = output.getvalue()[14:]
        else:
            raise ValueError('param (dtype) only "text" or "image" supported')
        wc.OpenClipboard()
        wc.EmptyClipboard()
        wc.SetClipboardData(type_data, data)
        wc.CloseClipboard()

    def Screenshot(hwnd, to_clipboard=True):
        '''为句柄为hwnd的窗口程序截图
        hwnd : 句柄
        to_clipboard : 是否复制到剪贴板
        '''
        import pyscreenshot as shot
        bbox = win32gui.GetWindowRect(hwnd)
        win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST, 0, 0, 0, 0,\
                              win32con.SWP_SHOWWINDOW|win32con.SWP_NOMOVE|win32con.SWP_NOSIZE)
        win32gui.SetWindowPos(hwnd, win32con.HWND_NOTOPMOST, 0, 0, 0, 0,\
                              win32con.SWP_SHOWWINDOW|win32con.SWP_NOMOVE|win32con.SWP_NOSIZE)
        win32gui.BringWindowToTop(hwnd)
        im = shot.grab(bbox)
        if to_clipboard:
            WxUtils.SetClipboard(im, 'image')
        return im
    
    def SavePic(savepath=None, filename=None):
        Pic = auto.WindowControl(ClassName='ImagePreviewWnd', Name='图片查看')
        Pic.SendKeys('{Ctrl}s')
        SaveAs = Pic.WindowControl(ClassName='#32770', Name='另存为...')
        SaveAsEdit = SaveAs.EditControl(ClassName='Edit', Name='文件名:')
        SaveButton = Pic.ButtonControl(ClassName='Button', Name='保存(S)')
        PicName, Ex = os.path.splitext(SaveAsEdit.GetValuePattern().Value)
        if not savepath:
            savepath = os.getcwd()
        if not filename:
            filename = PicName
        FilePath = os.path.realpath(os.path.join(savepath, filename + Ex))
        SaveAsEdit.SendKeys(FilePath)
        SaveButton.Click()
        Pic.SendKeys('{Esc}')

    def ControlSize(control):
        locate = control.BoundingRectangle
        size = (locate.width(), locate.height())
        return size
    
    def ClipboardFormats(unit=0, *units):
        units = list(units)
        wc.OpenClipboard()
        u = wc.EnumClipboardFormats(unit)
        wc.CloseClipboard()
        units.append(u)
        if u:
            units = WxUtils.ClipboardFormats(u, *units)
        return units

    def CopyDict():
        Dict = {}
        for i in WxUtils.ClipboardFormats():
            if i == 0:
                continue
            wc.OpenClipboard()
            try:
                content = wc.GetClipboardData(i)
                wc.CloseClipboard()
            except Exception as e:
                wc.CloseClipboard()
                raise ValueError
            if len(str(i))>=4:
                Dict[str(i)] = content
        return Dict


class WeChat():
    def __init__(self,Name,ClassName):

        self.UiaAPI = auto.WindowControl(Name=Name, ClassName=ClassName)
        self.UiaAPI.SetActive()
        # 切换窗口
        self.UiaAPI.SwitchToThisWindow()

        self.SessionList = self.UiaAPI.ListControl(Name='会话')
        self.EditMsg = self.UiaAPI.EditControl(Name='输入')
        self.SearchBox = self.UiaAPI.EditControl(Name='搜索')
        self.MsgList = self.UiaAPI.ListControl(Name='消息')
        
        self.SessionItemList = []

    def SendMsg(self, msg, clear=True):
        '''向当前窗口发送消息
        msg : 要发送的消息
        clear : 是否清除当前已编辑内容
        '''
        self.UiaAPI.SwitchToThisWindow()
        if clear:
            self.EditMsg.SendKeys('{Ctrl}a', waitTime=0)
        self.EditMsg.SendKeys(msg, waitTime=0)
        self.EditMsg.SendKeys('{Enter}', waitTime=0)

    def Search(self, keyword):
        '''
        查找微信好友或关键词
        keywords: 要查找的关键词，str   * 最好完整匹配，不完全匹配只会选取搜索框第一个
        '''
        self.UiaAPI.SetFocus()
        time.sleep(0.2)
        self.UiaAPI.SendKeys('{Ctrl}f', waitTime=1)
        self.SearchBox.SendKeys(keyword, waitTime=1.5)
        self.SearchBox.SendKeys('{Enter}')

    def ChatWith(self, who, RollTimes=None):
        '''
        打开某个聊天框
        who : 要打开的聊天框好友名，str;  * 最好完整匹配，不完全匹配只会选取搜索框第一个
        RollTimes : 默认向下滚动多少次，再进行搜索
        '''
        self.UiaAPI.SwitchToThisWindow()
        RollTimes = 10 if not RollTimes else RollTimes
        def roll_to(who=who, RollTimes=RollTimes):
            for i in range(RollTimes):
                if who not in self.GetSessionList()[:-1]:
                    self.SessionList.WheelDown(wheelTimes=3, waitTime=0.1*i)
                else:
                    time.sleep(0.5)
                    self.SessionList.ListItemControl(Name=who).Click(simulateMove=False)
                    return 1
            return 0
        rollresult = roll_to()
        if rollresult:
            return 1
        else:
            self.Search(who)
            return roll_to(RollTimes=1)

    def GetSessionList(self, reset=False):
        '''获取当前会话列表，更新会话列表'''
        self.SessionItem = self.SessionList.ListItemControl()
        SessionList = []
        if reset:
            self.SessionItemList = []
        for i in range(100):
            try:
                name = self.SessionItem.Name
            except:
                break
            if name not in self.SessionItemList:
                self.SessionItemList.append(name)
            if name not in SessionList:
                SessionList.append(name)
            self.SessionItem = self.SessionItem.GetNextSiblingControl()
        return SessionList

    def GetLastMessage(self):
        '''获取当前窗口中最后一条聊天记录'''
        auto.SetGlobalSearchTimeout(1.0)
        MsgItem = self.MsgList.GetChildren()[-1]
        Msg = WxUtils.SplitMessage(MsgItem)
        auto.SetGlobalSearchTimeout(10.0)
        return Msg
    
    def LoadMoreMessage(self, n=0.1):
        '''定位到当前聊天页面，并往上滚动鼠标滚轮，加载更多聊天记录到内存'''
        n = 0.1 if n<0.1 else 1 if n>1 else n
        self.MsgList.WheelUp(wheelTimes=int(500*n), waitTime=0.1)

    def GetAllMessage(self):
        '''获取当前窗口中加载的所有聊天记录'''
        MsgDocker = []
        MsgItems = self.MsgList.GetChildren()
        for MsgItem in MsgItems:
            MsgDocker.append(WxUtils.SplitMessage(MsgItem))
        return MsgDocker

    def GetAllText(self,MsgItems):
        if(MsgItems.LocalizedControlType != '文本'):
            MsgItem = MsgItems.GetChildren()
            for each in MsgItem:
                self.GetAllText(each)
        else:
            self.AllTextMsg.append(MsgItems.Name)

    def 获取聊天记录(self,MsgItems):
        if(MsgItems.LocalizedControlType != '编辑'):
            MsgItem = MsgItems.GetChildren()
            for each in MsgItem:
                self.获取聊天记录(each)
        else:
            if(not(MsgItems.Name in self.聊天记录)):
                self.聊天记录.append(MsgItems.Name)

    def GetAllTextItems(self,MsgItems):
        if(MsgItems.LocalizedControlType != '文本'):
            MsgItem = MsgItems.GetChildren()
            for each in MsgItem:
                self.GetAllTextItems(each)
        else:
            self.AllTextMsg.append(MsgItems)

    def GetAllMergeMessage(self,Partition):
        '''获取当前窗口中加载的所有合并的聊天记录'''
        Msg_list = []

        # 获取消息列表的所有子集
        MsgItems = self.MsgList.GetChildren()

        # 删除最后一次分割词以上的内容
        Partition_index = 0
        for i in range(len(MsgItems)):
            if(MsgItems[i].Name == Partition):
                Partition_index = i
        MsgItems = MsgItems[Partition_index:]

        # 获取所有的合并的聊天记录
        for MsgItem in MsgItems:
            if(MsgItem.Name == '[聊天记录]'):
                try:
                    self.AllTextMsg = []
                    self.GetAllTextItems(MsgItem)
                    位置 = self.AllTextMsg[0].BoundingRectangle
                    消息框 = self.MsgList.BoundingRectangle
                    while((位置.top<消息框.top) or (位置.bottom>消息框.bottom)):
                        if(位置.top<消息框.top):
                            self.MsgList.WheelUp(wheelTimes=5, waitTime=0.1)
                        if(位置.bottom>消息框.bottom):
                            self.MsgList.WheelDown(wheelTimes=5, waitTime=0.1)
                        位置 = self.AllTextMsg[0].BoundingRectangle
                        消息框 = self.MsgList.BoundingRectangle
                        print(位置.top,消息框.top,位置.bottom,消息框.bottom)

                    self.AllTextMsg[0].Click(waitTime=0.1)
                    # 获得聊天记录实例
                    self.聊天记录窗口 = auto.WindowControl(Name=self.AllTextMsg[0].Name, ClassName='ChatRecordWnd')
                    self.聊天记录窗口.SetActive()
                    # 切换窗口
                    self.聊天记录窗口.SwitchToThisWindow()
                
                    聊天记录列表 = self.聊天记录窗口.ListControl(LocalizedControlType='列表')
                    self.聊天记录 = []
                    for i in range(10):
                        self.获取聊天记录(聊天记录列表)
                        聊天记录列表.WheelDown(wheelTimes=int(5), waitTime=0.1)

                    Msg_list.append(self.聊天记录)
                    self.聊天记录窗口.SendKeys('{Esc}', waitTime=0)
                except:
                    pass
                # 切换窗口
                self.UiaAPI.SwitchToThisWindow()
        return Msg_list

def main():
    # 初始化一个ini文件进行数据存储
    dis = dispose_ini("config.ini")
    # 表格信息
    表格名称 = '微信信息导出表格.xlsx'
    try:
        dis.add_section("GroupChatList")
        dis.add_section("GroupChatListPartition")
        dis.add_section("config")
        print("检测到用户为第一次运行本程序，开始程序初始化流程:")
        print("请逐个输入需要提取信息的群聊名称，输入\'返回\'后结束本流程:")
        i = 0
        while(True):
            i += 1
            Group = input('请输入第'+str(i)+'个群名称:')
            if(Group != '返回'):
                dis.set_option("GroupChatList","GroupChat"+str(i),Group)
                dis.set_option("GroupChatListPartition","GroupChat"+str(i)+'Partition','')
            else:
                break
        print("初始化内容结束")
    except:
        pass
    
    while(True):
        print("\n请确定当前需要进行的任务内容(请输入:1或者2):\n1.微信自动化获取群消息。\n2.淘宝导出表格的价格导入\n")
        option = input("请输入:")
        if(option == '1'):
            try:
                GroupChatList = dis.get_options('GroupChatList')
                # 获得微信实例
                wx = WeChat(Name='微信', ClassName='WeChatMainWndForPC')
                # 已处理过的信息
                ProcessedMsg = []
                for wxql in GroupChatList:
                    # 寻找到相应的群聊并点击
                    QlName = dis.get_option('GroupChatList',wxql)
                    wx.Search(QlName)

                    # 获取分割词，若没有则添加
                    Partition = dis.get_option('GroupChatListPartition',wxql+'Partition')
                    if(Partition == ''):
                        Partition = QlName+'>>分割线'
                        wx.SendMsg(Partition)
                        dis.set_option("GroupChatListPartition",wxql+'Partition',Partition)
                        print('分割信息添加完成，该群聊初始化已完成，后续请勿删除该群聊信息。')
                        continue
                
                    LoadMoreNum = 0
                    # 寻找分隔词
                    while(True):
                        LoadMoreNum += 1
                        print('正在寻找:'+Partition)
                        AllMsg = wx.GetAllMessage()
                        for each in AllMsg:
                            if each[1] == Partition:
                                LoadMoreNum = 100
                                break
                        if(LoadMoreNum >= 100):
                            break
                        wx.LoadMoreMessage()

                    # 获取合并的聊天记录的消息内容
                    Msg_list = wx.GetAllMergeMessage(Partition)
                    
                    # 对消息内容进行处理
                    for each in Msg_list:
                        # 创建一个空的,长度为15，内容为''的列表
                        MsgData = []
                        for i in range(15):
                            MsgData.append('')

                        if(not(each in ProcessedMsg)):
                            try:
                                msg_data = each[0].split('\n')
                            except:
                                pass

                            try:
                                msg_data[2] = re.sub(' +', ' ', msg_data[2].upper())
                                msg_data_c = msg_data[2].split(' ')
                                货号与尺码 = True
                            except:
                                货号与尺码 = False
                            try:
                                MsgData[0] = msg_data[1]
                            except:
                                print('订单编号写入错误')

                            try:
                                if(货号与尺码):
                                    MsgData[2] = msg_data_c[0]
                                    if(len(msg_data_c) > 1):
                                        MsgData[3] = msg_data_c[1]
                            except:
                                print('货号与尺码写入错误')

                            try:
                                MsgData[4] = msg_data[3]
                            except:
                                print('调货价格写入错误')

                            try:
                                index = msg_data[0].find(' : ')
                                if(index != -1):
                                    MsgData[1] = msg_data[0][0:index]
                            except:
                                print('调货员写入错误')

                            try:
                                for i in range(1,len(each)):
                                    if(each[i].find('zfb')!=-1):
                                        支付宝信息序列号 = i
                                        each[支付宝信息序列号] = each[支付宝信息序列号].replace(',',' ')
                                        each[支付宝信息序列号] = each[支付宝信息序列号].replace('，',' ')
                                        each[支付宝信息序列号] = re.sub(' +', ' ', each[支付宝信息序列号])
                                        支付宝信息 = each[支付宝信息序列号].split(' ')
                                        # 如果第二位是手机号或者邮箱，则为账号
                                        if(len(支付宝信息)>=3):
                                            if(re.match(r"^(13[0-9]|14[01456879]|15[0-35-9]|16[2567]|17[0-8]|18[0-9]|19[0-35-9])\d{8}$", 支付宝信息[1]) or re.match(r"^\w+([-+.]\w+)*@\w+([-.]\w+)*\.\w+([-.]\w+)*$", 支付宝信息[1])):
                                                MsgData[13] = 支付宝信息[1]
                                                MsgData[14] = 支付宝信息[2]
                                            else:
                                                MsgData[13] = 支付宝信息[2]
                                                MsgData[14] = 支付宝信息[1]
                            except:
                                print('支付宝信息写入错误')

                            write_excel_xlsx_append(表格名称, [MsgData,])
                
                    wx.SendMsg('群聊'+QlName+'的消息已处理')
                    print('群聊'+QlName+'的消息已处理')
                
                    Partition = QlName+'>>分割线'
                    wx.SendMsg(Partition)
                    dis.set_option("GroupChatListPartition",wxql+'Partition',Partition)
            except Exception as e:
                # 输出错误的具体原因
                print(e)
                print(sys.exc_info())

                # 输出错误的具体位置，报错行号位置在第几行
                print('\n', '>>>' * 20)
                print(traceback.print_exc())
                print('\n', '>>>' * 20)
                print(traceback.format_exc())

        elif(option == '2'):
            try:
                print('请选择需要处理的淘宝导出表格:')
                # 获取淘宝所有订单的价格
                try:
                    FileName = easygui.fileopenbox()
                except Exception as e:
                    print('表格操作错误')
                    # 输出错误的具体原因
                    print(e)
                    print(sys.exc_info())

                    # 输出错误的具体位置，报错行号位置在第几行
                    print('\n', '>>>' * 20)
                    print(traceback.print_exc())
                    print('\n', '>>>' * 20)
                    print(traceback.format_exc())

                taobaolist = []
                with open(FileName, 'r', encoding='gbk',errors='ignore') as csv_f:
                    reader = csv.DictReader(csv_f, skipinitialspace=True)
                    for cvs_row in reader:
                        #if(cvs_row['宝贝标题 ' or '宝贝标题'].find('买家秀征集赢奖励！！！ ~ 收货联系客服') == -1):
                        taobaolist.append({
                                '订单编号':re.sub(u"([^\u0030-\u0039])", "", cvs_row['订单编号']),
                                '价格':cvs_row['买家实际支付金额'],
                                '备注':cvs_row['订单备注'],
                                '店铺名称':cvs_row['店铺名称'],
                                '邮费':cvs_row['买家应付邮费']
                                })
            
                # 获取微信表格的所有订单编号
                dqworkbook = openpyxl.load_workbook(表格名称)
                dqsheet = dqworkbook.worksheets[0]
                wxlist = []
                for row in dqsheet.rows:
                    if(row[0].value != None):
                        wxlist.append(re.sub(u"([^\u0030-\u0039])", "", row[0].value))

                # 打开表格，写入数据
                xrworkbook = openpyxl.load_workbook(表格名称)
                xrsheet = xrworkbook.worksheets[0]
                for i in range(len(wxlist)):
                    for each in taobaolist:
                        if(wxlist[i] == each['订单编号']):
                            xrsheet.cell(row=i+1, column=6, value=each['价格'])
                            xrsheet.cell(row=i+1, column=7, value=each['备注'])
                            xrsheet.cell(row=i+1, column=12, value=each['邮费'])
                            xrsheet.cell(row=i+1, column=13, value=each['店铺名称'])
                            
                try:
                    xrworkbook.save(表格名称)
                    print('淘宝数据处理完成\n')
                except:
                    print('淘宝数据保存失败\n')
            
            except Exception as e:
                # 输出错误的具体原因
                print(e)
                print(sys.exc_info())

                # 输出错误的具体位置，报错行号位置在第几行
                print('\n', '>>>' * 20)
                print(traceback.print_exc())
                print('\n', '>>>' * 20)
                print(traceback.format_exc())

        else:
            print('错误的命令，请重新输入')
if __name__ == "__main__":
    main()
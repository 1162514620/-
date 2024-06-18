import time
import sys
import os
import keyboard
import threading
import uiautomation as auto
import openpyxl
import easygui
import re
import csv
import traceback

from 配置文件处理 import 配置文件处理

表格头 = ['订单编号', '调货员/现货', '货号', '尺码', '调货价格', '淘宝价格', '商家备忘',
       '点数', '扣除', '利润', '备注', '运费', '店铺名称', '支付宝', '昵称']

def 获取表格实例(path, sheet_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.save(path)

def 追加方式写入表格(path, value, truncate_sheet=False):
    if not os.path.exists(path):
        获取表格实例(path, 'Sheet1')
        data = openpyxl.load_workbook(path)
        # 取第一张表
        sheetnames = data.sheetnames
        sheet = data[sheetnames[0]]
        for i in range(0, len(表格头)):
            sheet.cell(row=1, column=i + 1, value=str(表格头[i]))
        data.save(path)
    data = openpyxl.load_workbook(path)
    # 取第一张表
    sheetnames = data.sheetnames
    sheet = data[sheetnames[0]]
    sheet = data.active
    if (truncate_sheet):  # truncate_sheet为True，覆盖原表中的数据
        startrows = 0
    else:
        startrows = sheet.max_row  # 获得行数
    index = len(value)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=startrows + i + 1, column=j +
                       1, value=str(value[i][j]))
    data.save(path)

def 读取表格(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

class 微信自动化():
    def __init__(self):
        # 获取微信窗口
        self.微信窗口 = auto.WindowControl(Name='微信', ClassName='WeChatMainWndForPC')
        # 设置窗口为活跃状态，放到最顶层
        self.微信窗口.SetActive()
        # 切换到微信窗口
        self.微信窗口.SwitchToThisWindow()

        # 一些常用的控件筛选
        self.消息列表 = self.微信窗口.ListControl(Name='消息')

    def 发送消息(self, Name, msg):
        '''向当前窗口发送消息
        Name : 需要发送消息的群聊
        msg : 要发送的消息
        clear : 是否清除当前已编辑内容
        '''
        self.EditMsg = self.微信窗口.EditControl(Name=Name)
        self.微信窗口.SwitchToThisWindow()
        self.EditMsg.SendKeys('{Ctrl}a')
        self.EditMsg.SendKeys(msg)
        self.EditMsg.SendKeys('{Enter}')

    def 查找好友(self, keyword):
        '''
        查找微信好友或关键词
        keywords: 要查找的关键词，str   * 最好完整匹配，不完全匹配只会选取搜索框第一个
        '''
        self.SearchBox = self.微信窗口.EditControl(Name='搜索')
        self.微信窗口.SetFocus()
        time.sleep(0.2)
        self.微信窗口.SendKeys('{Ctrl}f', waitTime=1)
        self.SearchBox.SendKeys(keyword, waitTime=1.5)
        self.SearchBox.SendKeys('{Enter}')

    def 获取更多消息(self, n=0.1):
        '''定位到当前聊天页面，并往上滚动鼠标滚轮，加载更多聊天记录到内存'''
        n = 0.1 if n < 0.1 else 1 if n > 1 else n
        self.消息列表.WheelUp(wheelTimes=int(50*n))

    def 获取合并消息的内容(self, msg):
        位置 = msg.BoundingRectangle
        消息框 = self.消息列表.BoundingRectangle
        # 寻找到正确的消息位置
        try:
            # 只要合并消息对象按钮在消息框的外面，就一直寻找
            while ((位置.top < 消息框.top) or (位置.bottom > 消息框.bottom)):
                if (位置.top < 消息框.top):
                    self.消息列表.WheelUp(wheelTimes=1, waitTime=0.1)
                if (位置.bottom > 消息框.bottom):
                    self.消息列表.WheelDown(wheelTimes=1, waitTime=0.1)
                位置 = msg.BoundingRectangle
            msg.ButtonControl(Name = "").Click(waitTime=0.1)
        except Exception as e:
            # 输出错误的具体位置，报错行号位置在第几行
            print('\n', '>>>' * 20)
            print(traceback.print_exc())
            print("定位合并消息错误")

        try:
            # 打开并激活切换到合并消息窗口
            合并消息窗口 = auto.WindowControl(ClassName='ChatRecordWnd')
            合并消息窗口.SetActive()
            合并消息窗口.SwitchToThisWindow()

            全部消息列表 = []
            返回消息列表 = []
            记录次数 = 0
            记录消息长度 = 0
            for i in range(50):

                当前页消息对象列表 = 合并消息窗口.ListControl(Name = "消息记录").GetChildren()
                for 当前页消息对象 in 当前页消息对象列表:
                    if  当前页消息对象.Name not in 全部消息列表:
                        全部消息列表.append(当前页消息对象.Name)
                        
                        if "[图片]" not in 当前页消息对象.Name:
                            返回消息列表.append(当前页消息对象.TextControl(foundIndex = 1).Name)
                            返回消息列表.append(当前页消息对象.TextControl(foundIndex = 3).Name)

                if (记录消息长度 == len(全部消息列表)): 
                    记录次数 += 1
                    if (记录次数 >= 8):
                        break
                else:
                    记录次数 = 0
                    记录消息长度 = len(全部消息列表)

                合并消息窗口.WheelDown(wheelTimes=int(5), waitTime=0.1)
                time.sleep(0.1)
            合并消息窗口.SendKeys('{Esc}', waitTime=0)
        except Exception as e:
            合并消息窗口.SendKeys('{Esc}', waitTime=0)
            # 输出错误的具体位置，报错行号位置在第几行
            print('\n', '>>>' * 20)
            print(traceback.print_exc())
            print("聊天记录获取错误")
        
        # 设置窗口为活跃状态，放到最顶层
        self.微信窗口.SetActive()
        # 切换到微信窗口
        self.微信窗口.SwitchToThisWindow()

        return 返回消息列表
    
def main():
    # 初始化一个ini文件进行数据存储
    配置信息 = 配置文件处理("配置信息.ini")
    # 表格信息
    表格名称 = '微信信息导出表格.xlsx'
    try:
        配置信息.add_section("群聊列表")
        配置信息.add_section("群聊分割")
        配置信息.add_section("config")
        print("检测到用户为第一次运行本程序，开始程序初始化流程:")
        print("请逐个输入需要提取信息的群聊名称，输入\'返回\'后结束本流程:")
        i = 0
        while (True):
            i += 1
            群聊名称 = input('请输入第'+str(i)+'个群名称:')
            if (群聊名称 != '返回'):
                配置信息.set_option("群聊列表", "群聊"+str(i), 群聊名称)
                配置信息.set_option("群聊分割",
                               "群聊"+str(i)+'分割内容', '')
            else:
                break
        print("初始化内容结束")
    except:
        pass

    while (True):
        print("\n请确定当前需要进行的任务内容(请输入:1或者2):\n1.微信自动化获取群消息。\n2.淘宝导出表格的价格导入\n")
        option = input("请输入:")
        if (option == '1'):
            try:
                群聊列表 = 配置信息.get_options('群聊列表')
                # 获得微信实例
                微信实例 = 微信自动化()
                for 微信群聊 in 群聊列表:
                    # 寻找到相应的群聊并点击
                    群聊名称 = 配置信息.get_option('群聊列表', 微信群聊)
                    微信实例.查找好友(群聊名称)
                    # 获取分割词，若没有则添加
                    分割内容 = 配置信息.get_option('群聊分割', 微信群聊+'分割内容')
                    if (分割内容 == ''):
                        分割内容 = 群聊名称+'>>分割线'
                        微信实例.发送消息(群聊名称, 分割内容)
                        配置信息.set_option("群聊分割", 微信群聊+'分割内容', 分割内容)
                        print('分割信息添加完成，该群聊初始化已完成，后续请勿删除该群聊信息。')
                        continue

                    
                    滚动次数 = 0
                    分割标志 = None
                    # 寻找分隔词，并获取合并的消息列表对象
                    while (滚动次数 < 100):
                        滚动次数 += 1
                        print('正在寻找:'+分割内容)
                        # 获取当前对话窗口的所有消息对象的列表
                        消息对象列表 = 微信实例.消息列表.GetChildren()
                        for 消息对象 in 消息对象列表:
                            if 消息对象.Name == 分割内容:
                                滚动次数 = 100
                                分割标志 = 消息对象.GetRuntimeId()
                        
                        微信实例.获取更多消息()
                                
                    
                    所有合并消息列表 = []
                    # 刷新消息列表
                    消息对象列表 = 微信实例.消息列表.GetChildren()   
                    # 切割消息列表 
                    for i, 消息对象 in enumerate(消息对象列表):
                        if 消息对象.GetRuntimeId() == 分割标志:
                            消息对象列表 = 消息对象列表[i:]

                    # 对消息对象列表进行一次排序
                    消息对象列表.sort(key=lambda item: item.BoundingRectangle.top)

                    # 获取后面的所有聊天记录
                    for 消息对象 in 消息对象列表:
                        if 消息对象.Name == "[聊天记录]":
                            if 消息对象 not in 所有合并消息列表:
                                所有合并消息列表.append(消息对象)

                    合并消息文本列表 = []
                    # 获取所有合并消息对象的文本内容
                    for 合并的消息 in 所有合并消息列表:
                        消息数据 = 微信实例.获取合并消息的内容(合并的消息)
                        if 消息数据 not in 合并消息文本列表:
                            合并消息文本列表.append(消息数据)

                    已处理过的信息 = []
                    # 对消息内容进行处理
                    for 合并消息文本 in 合并消息文本列表:
                        # 创建一个空的,长度为15，内容为''的列表
                        写入数据 = []
                        for i in range(15):
                            写入数据.append('')

                        if  消息数据 not in 已处理过的信息:
                            # 分割订单的基本信息，地址，订单号，货号与尺码，金额，昵称
                            try:
                                订单基本信息 = 合并消息文本[1].split('\n')

                                订单地址 = 订单基本信息[0]
                                订单号 = 订单基本信息[1]
                                订单的货号与尺码 = 订单基本信息[2]
                                调货价格 = 订单基本信息[3]
                            except:
                                print('订单信息分割错误')
                            
                            # 订单号的写入
                            写入数据[0] = 订单号
                            # 调货员的写入
                            写入数据[1] = 合并消息文本[0]
                            # 调货价格的写入
                            写入数据[4] = 调货价格
                            

                            # 货号与尺码的写入
                            try:
                                订单的货号与尺码 = re.sub(' +', ' ', 订单的货号与尺码.upper())
                                货号,尺码 = 订单的货号与尺码.split(' ')
                                写入数据[2] = 货号
                                if (尺码 != None):
                                    写入数据[3] = 尺码
                            except:
                                print('货号与尺码写入错误')

                            try:
                                for i in range(1, len(合并消息文本)):
                                    if (合并消息文本[i].find('zfb') != -1):
                                        支付宝信息序列号 = i
                                        合并消息文本[支付宝信息序列号] = 合并消息文本[支付宝信息序列号].replace(',', ' ')
                                        合并消息文本[支付宝信息序列号] = 合并消息文本[支付宝信息序列号].replace('，', ' ')
                                        合并消息文本[支付宝信息序列号] = re.sub(' +', ' ', 合并消息文本[支付宝信息序列号])
                                        支付宝信息 = 合并消息文本[支付宝信息序列号].split(' ')
                                        # 如果第二位是手机号或者邮箱，则为账号
                                        if (len(支付宝信息) >= 3):
                                            if (re.match(r"^(13[0-9]|14[01456879]|15[0-35-9]|16[2567]|17[0-8]|18[0-9]|19[0-35-9])\d{8}$", 支付宝信息[1]) or re.match(r"^\w+([-+.]\w+)*@\w+([-.]\w+)*\.\w+([-.]\w+)*$", 支付宝信息[1])):
                                                写入数据[13] = 支付宝信息[1]
                                                写入数据[14] = 支付宝信息[2]
                                            else:
                                                写入数据[13] = 支付宝信息[2]
                                                写入数据[14] = 支付宝信息[1]
                            except:
                                print('支付宝信息写入错误')

                            追加方式写入表格(表格名称, [写入数据,])

                    微信实例.发送消息(群聊名称, '群聊'+群聊名称+'的消息已处理')
                    print('群聊'+群聊名称+'的消息已处理')

                    分割内容 = 群聊名称+'>>分割线'
                    微信实例.发送消息(群聊名称, 分割内容)
                    配置信息.set_option("群聊分割",
                                   微信群聊+'分割内容', 分割内容)
            except Exception as e:
                # 输出错误的具体位置，报错行号位置在第几行
                print('\n', '>>>' * 20)
                print(traceback.print_exc())

        elif (option == '2'):
            try:
                print('请选择需要处理的淘宝导出表格:')
                # 获取淘宝所有订单的价格
                try:
                    FileName = easygui.fileopenbox()
                except:
                    print('打开表格错误\n')

                taobaolist = []
                with open(FileName, 'r', encoding='gbk', errors='ignore') as csv_f:
                    reader = csv.DictReader(csv_f, skipinitialspace=True)
                    for cvs_row in reader:
                        # if(cvs_row['宝贝标题 ' or '宝贝标题'].find('买家秀征集赢奖励！！！ ~ 收货联系客服') == -1):
                        taobaolist.append({
                            '订单编号': re.sub(u"([^\u0030-\u0039])", "", cvs_row['订单编号']),
                            '价格': cvs_row['买家实际支付金额'],
                            '备注': cvs_row['商家备注'],
                            '店铺名称': cvs_row['店铺名称'],
                            '邮费': cvs_row['买家应付邮费']
                        })
 
                # 获取微信表格的所有订单编号
                dqworkbook = openpyxl.load_workbook(表格名称)
                dqsheet = dqworkbook.worksheets[0]
                wxlist = []
                for row in dqsheet.rows:
                    if (row[0].value != None):
                        wxlist.append(
                            re.sub(u"([^\u0030-\u0039])", "", row[0].value))

                # 打开表格，写入数据
                xrworkbook = openpyxl.load_workbook(表格名称)
                xrsheet = xrworkbook.worksheets[0]
                for i in range(len(wxlist)):
                    for each in taobaolist:
                        if (wxlist[i] == each['订单编号']):
                            xrsheet.cell(row=i+1, column=6, value=each['价格'])
                            xrsheet.cell(row=i+1, column=7, value=each['店铺名称'])
                            xrsheet.cell(row=i+1, column=11, value=each['备注'])
                            xrsheet.cell(row=i+1, column=12, value=each['邮费'])
                            xrsheet.cell(row=i+1, column=13, value=each['店铺名称'])

                try:
                    xrworkbook.save(表格名称)
                    print('淘宝数据处理完成\n')
                except:
                    print('淘宝数据保存失败\n')

            except Exception as e:
                # 输出错误的具体位置，报错行号位置在第几行
                print('\n', '>>>' * 20)
                print(traceback.print_exc())

        else:
            print('错误的命令，请重新输入')


def 按键监听():
    keyboard.wait('F12')
    os._exit(0)


if __name__ == "__main__":
    print('欢迎使用!')
    threading.Thread(target=按键监听).start()
    main()
